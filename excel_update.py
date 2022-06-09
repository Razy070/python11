import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# чтение
# обрабатываем данные
# записываем в исходный файл (можно исходный файл удалять)


workbook = openpyxl.load_workbook("""temp/sample_example.xlsx""")

# берём активную страницу из рабочей книги
worksheet = workbook.active

# берём последнюю строку в excel-файле
max_row = worksheet.max_row
print(max_row)
# берём последнюю колонку в excel-файле
max_column = worksheet.max_column
print(max_column)

# записать всё в массив с массивами [["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"], ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]...]
external_array = []

# наполнение внешнего массива
for row in range(1, max_row + 1):
    internal_array = []

    # наполнение внутреннего массива
    for col in range(1, max_column + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is None:
            value = ''
            # continue  # (продолжить) оператор позволяет пропустить эту итерацию цикла
            # break  # (прекратить) оператор позволяет полностью остановить цикл
        internal_array.append(value)
    # завершение наполнение внутреннего массива

    if len(internal_array) < 1:
        continue

    external_array.append(internal_array)
# завершение наполнение внешнего массива

print(external_array)

# index = 0
# while True:
#     index += 1
#     if index > 50:
#         break # (прекратить) оператор позволяет полностью остановить цикл

workbook_new = Workbook()
worksheet_new = workbook_new.active

# col_count = external_array[0]
print(
    f"col_count: {external_array}")  # [['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', ''], ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', ''],]
print(f"col_count: {external_array[0]}")  # ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']
print(f"col_count: {external_array[0][1]}")  # 'Almaty'
print(f"col_count: {external_array[0][1][2:-2:1]}")  # 'ma'


def function_len_array(array):  # ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']
    return len(array)  # 5


# -2 -1 0 1 2

for row in range(0, function_len_array(external_array)):  # [0, 1, 2, 3, 4, ..., 1007]
    # print(f"col_count: {len(external_array[row])}")
    for col in range(0, function_len_array(external_array[row])):
        # [0, 1, 2, 3, 4, 5]  # external_array[row] == ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']

        worksheet[f'{get_column_letter(col + 1)}{row + 1}'] = external_array[row][col]
        if row == 0:
            worksheet[f'{get_column_letter(col + 1)}{row + 1}'].font = Font(bold=True)
        pass
    # print(external_array[row])
    pass

# for char in 'ABCDE':
# worksheet_new['A1'].font = Font(bold=True)
# worksheet_new['B1'].font = Font(bold=True)
# worksheet_new['C1'].font = Font(bold=True)
# worksheet_new['D1'].font = Font(bold=True)
# worksheet_new['E1'].font = Font(bold=True)
ws['A1'] = "а_1"

workbook.save('temp/sample_example_new.xlsx')

wb = Workbook()
ws = wb.active
ws['A1'] = "а_1"
ws['B1'] = "б_1"
ws['C1'] = "ц_1"
ws['A2'] = "а_2"
ws['B2'] = "б_2"
ws['C2'] = "ц_2"
ws['A3'] = "а_3"
ws['B3'] = "б_3"
ws['C3'] = "ц_3"
ws['A4'] = "а_4"
ws['B4'] = "б_4"
ws['C4'] = "ц_4"
ws['A5'] = "а_5"
ws['B5'] = "б_5"
ws['C5'] = "ц_5"
ws['A6'] = "а_6"
ws['B6'] = "б_6"
ws['C6'] = "ц_6"
ws['A7'] = "а_7"
ws['B7'] = "б_7"
ws['C7'] = "ц_7"
ws['A8'] = "а_8"
ws['B8'] = "б_8"
ws['C8'] = "ц_8"
ws['A9'] = "а_9"
ws['B9'] = "б_9"
ws['C9'] = "ц_9"
ws['A10'] = "а_10"
ws['B10'] = "б_10"
ws['C10'] = "ц_10"
ws['A11'] = "а_11"
ws['B11'] = "б_11"
ws['C11'] = "ц_11"
ws['A12'] = "а_12"
ws['B12'] = "б_12"
ws['C12'] = "ц_12"
ws['A13'] = "а_13"
ws['B13'] = "б_13"
ws['C13'] = "ц_13"
ws['A14'] = "а_14"
ws['B14'] = "б_14"
ws['C14'] = "ц_14"
ws['A15'] = "а_15"
ws['B15'] = "б_15"
ws['C15'] = "ц_15"
ws['A16'] = "а_16"
ws['B16'] = "б_16"
ws['C16'] = "ц_16"
ws['A17'] = "а_17"
ws['B17'] = "б_17"
ws['C17'] = "ц_17"
ws['A18'] = "а_18"
ws['B18'] = "б_18"
ws['C18'] = "ц_18"
ws['A19'] = "а_19"
ws['B19'] = "б_19"
ws['C19'] = "ц_19"
ws['B20'] = "б_20"
ws['C20'] = "ц_20"
ws['B20'] = "б_21"
ws['C21'] = "ц_21"
ws['B21'] = "б_22"
ws['C21'] = "ц_22"
ws['B22'] = "б_22"
ws['C21'] = "ц_23"
ws['C22'] = "ц_24"
ws['C23'] = "ц_24"
ws['C24'] = "ц_24"
wb.save("temp/BoldDemo.xlsx")

# читаем первый файл, второй, третий... файлы
# складываем значение в каждом уникальном массиве
# создаём новый файл, который содержит в себе финальное значение (+ форматирование)
# как получить все имена файлов в папке

# форматирование: жирный, курсив, размер, capitalize, фон, границы, подчёркивание, формулы, объединение, подбор ширины
